import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import mysql.connector  # Import MySQL connector library
# from ..settings import DB_USER_MASA, DB_PASSWORD_MASA
import yaml

import sys
sys.path.append('./')
import utils.formula as formula
from db import database_utils
from sqlalchemy import create_engine
from utils.auxiliar_functions import use_semgrep, dekra_script_version, export_csv

# Replace with your actual database connection details
db_config = {
    "host": "localhost",
    "user": 'masa_script',
    "password": 'MASA123',
    "database": "automated_MASA"
}

TABLE_REPORT = 'Report'
TABLE_FAIL_COUNTS = 'Total_Fail_Counts'
TABLE_LOGGING = 'Logging'
TABLE_SEMGREP = 'SEMGREP_FINDINGS'
TABLE_DEKRA = 'DEKRA_FINDINGS'

actual_timestamp = sys.argv[1]
actual_date = actual_timestamp.split('_')[0]
actual_date_path = os.path.join(os.getcwd() + '/apks/Results', actual_date)
actual_hour = actual_timestamp.split('_')[1]
path_export_csv = os.path.join(os.getcwd() + '/apks/Results', actual_date, actual_hour)


def extract_data(table_name):
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT tn.* FROM " + table_name + " tn INNER JOIN Report r ON tn.HASH = r.HASH WHERE r.TIMESTAMP = '" + actual_timestamp + "'"
    df = pd.read_sql(query, engine)
    return df

def extract_report(table_name):
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT HASH, APP_NAME, VERSION_NAME, SEMGREP, SCRIPT_VERSION, CODE_1, CODE_2, CRYPTO_1, CRYPTO_3, NETWORK_1, NETWORK_2, NETWORK_3, PLATFORM_2, PLATFORM_3, STORAGE_2 FROM " + table_name + " WHERE TIMESTAMP = '" + actual_timestamp + "'"
    df = pd.read_sql(query, engine)
    return df

def collect_data_dekra(dekra, actual_timestamp):

    # Extract data from the database
    report_table = extract_report(TABLE_REPORT, )
    fail_counts_table = extract_data(TABLE_FAIL_COUNTS)
    logging_table = extract_data(TABLE_LOGGING)
    dekra_findings = extract_data(TABLE_DEKRA)

    # Formula execution and data extraction
    database_utils.unify_suid_permissions(actual_timestamp)

    # Obtain test list
    with open('config/methods_config.yml') as f:
        config = yaml.load(f, Loader=yaml.SafeLoader)

    tests_list = [value.upper() for values in config['tests'].values() for value in values]

    value = formula.calculate_formula(0.01, 0.01, tests_list, actual_timestamp)

    if export_csv():
        report_table.to_csv(path_export_csv + '/Preload_Analysis_Report_' + actual_timestamp + '.csv', index=False)
        fail_counts_table.to_csv(path_export_csv + '/Preload_Analysis_Total_Fail_Counts_' + actual_timestamp + '.csv', index=False)
        logging_table.to_csv(path_export_csv + '/Preload_Analysis_Logging_' + actual_timestamp + '.csv', index=False)
        dekra_findings.to_csv(path_export_csv + '/Preload_Analysis_Findings_' + actual_timestamp + '.csv', index=False)
        data_formula = [['Formula'], [value]]
        df = pd.DataFrame(data_formula)
        # Export the DataFrame to a CSV file with no index and no header
        df.to_csv(path_export_csv + '/Preload_Analysis_Formula_' + actual_timestamp + '.csv', index=False, header=False)

    else:
        # Create a new Excel workbook
        workbook = Workbook()
        report_sheet = workbook.create_sheet(title='Report')
        findings_sheet = workbook.create_sheet(title='Findings')
        fail_counts_sheet = workbook.create_sheet(title='Total_Fail_Counts')
        formula_sheet = workbook.create_sheet(title='Formula')
        logging_sheet = workbook.create_sheet(title='Logging')

        # Convert DataFrames to rows and write to the corresponding sheets
        for row in dataframe_to_rows(report_table, index=False, header=True):
            report_sheet.append(row)

        for row in dataframe_to_rows(dekra_findings, index=False, header=True):
            findings_sheet.append(row)

        for row in dataframe_to_rows(fail_counts_table, index=False, header=True):
            fail_counts_sheet.append(row)

        for row in dataframe_to_rows(logging_table, index=False, header=True):
            logging_sheet.append(row)

        formula_sheet.append(["Formula Value"])
        formula_sheet.append([value])

        # Remove the default empty sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        # Save the workbook to a file
        workbook.save(path_export_csv + '/Preload_Analysis_' + actual_timestamp + '.xlsx')

def collect_data_semgrep(dekra):
    report_table = extract_report(TABLE_REPORT)
    finding_table = extract_data(TABLE_SEMGREP)

    if export_csv():
        report_table.to_csv(path_export_csv + '/Preload_Analysis_Report_' + actual_timestamp + '.csv', index=False)
        finding_table.to_csv(path_export_csv + '/Preload_Analysis_Findings_' + actual_timestamp + '.csv', index=False)

    else:
        # Create a new Excel workbook
        workbook = Workbook()
        report_sheet = workbook.create_sheet(title='Report')
        report_findings = workbook.create_sheet(title='Findings')

        # Convert DataFrames to rows and write to the corresponding sheets
        for row in dataframe_to_rows(report_table, index=False, header=True):
            report_sheet.append(row)

        # Convert DataFrames to rows and write to the corresponding sheets
        for row in dataframe_to_rows(finding_table, index=False, header=True):
            report_findings.append(row)

        # Remove the default empty sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        # Save the workbook to a file
        workbook.save(path_export_csv + '/Preload_Analysis_' + actual_timestamp + '.xlsx')

if __name__ == "__main__":

    # Verify if some apps has been scanned with timestamp
    scanned_apks = database_utils.get_scanned_apks(actual_timestamp)

    if scanned_apks == 0:
        print('Could not export data because there were no applications to scan.')
    else:

        dekra = not use_semgrep()

        if not os.path.exists(actual_date_path):
            try:
                os.mkdir(actual_date_path)
            except Exception as e:
                print(f"Error creating {actual_date_path} folder: {e}")

        if not os.path.exists(path_export_csv):
            try:
                os.mkdir(path_export_csv)
            except Exception as e:
                print(f"Error creating {path_export_csv} folder: {e}")

        if dekra:
            collect_data_dekra(dekra, actual_timestamp)
        else:
            collect_data_semgrep(dekra)
