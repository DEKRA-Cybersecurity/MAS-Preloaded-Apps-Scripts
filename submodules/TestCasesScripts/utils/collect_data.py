import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import mysql.connector  # Import MySQL connector library
# from ..settings import DB_USER_MASA, DB_PASSWORD_MASA
import yaml

import sys
sys.path.append('./')
import utils.formula as formula
from db import database_utils
from sqlalchemy import create_engine
from utils.auxiliar_functions import use_semgrep, parse_timestamp, get_script_version, export_csv
from settings import DB_USER_MASA, DB_PASSWORD_MASA, DB_HOST_MASA

# Replace with your actual database connection details
db_config = {
    "host": DB_HOST_MASA,
    "user": DB_USER_MASA,
    "password": DB_PASSWORD_MASA,
    "database": database_utils.get_database_name()
}

TABLE_REPORT = 'Report'
TABLE_FAIL_COUNTS = 'Total_Fail_Counts'
TABLE_LOGGING = 'Logging'
TABLE_FINDINGS = 'Findings'
TABLE_PERMISSIONS = 'Permissions'

timestamp = sys.argv[1]
uuid_execution = sys.argv[2]
actual_timestamp = parse_timestamp(timestamp)
path_export_csv = os.path.join(os.getcwd() + '/Results', actual_timestamp + '_' +  uuid_execution)

def extract_findings():
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT f.HASH, f.APP_NAME, f.CATEGORY, f.CHECK_ID, f.PATH, f.LINE FROM Findings f INNER JOIN Report r ON f.HASH = r.HASH and f.ID_EXECUTION = r.ID_EXECUTION WHERE r.ID_EXECUTION = '" + uuid_execution + "'"
    df = pd.read_sql(query, engine)
    return df

def extract_total_fail_counts():
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT tfc.HASH, tfc.CODE_1, tfc.CODE_2, tfc.CRYPTO_1, tfc.CRYPTO_3, tfc.NETWORK_1, tfc.NETWORK_2, tfc.NETWORK_3, tfc.PLATFORM_2, tfc.PLATFORM_3, tfc.STORAGE_2 FROM Total_Fail_Counts tfc INNER JOIN Report r ON tfc.HASH = r.HASH AND tfc.ID_EXECUTION = r.ID_EXECUTION WHERE tfc.ID_EXECUTION = '" + uuid_execution + "'"
    df = pd.read_sql(query, engine)
    return df

def extract_report():
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT HASH, APP_NAME, VERSION_NAME, SEMGREP, SCRIPT_VERSION, CODE_1, CODE_2, CRYPTO_1, CRYPTO_3, NETWORK_1, NETWORK_2, NETWORK_3, PLATFORM_2, PLATFORM_3, STORAGE_2 FROM Report WHERE ID_EXECUTION = '" + uuid_execution + "'"
    df = pd.read_sql(query, engine)
    return df

def extract_logging():
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT id, HASH, APP_NAME, TIMESTAMP, TESTCASE, ERROR FROM Logging WHERE ID_EXECUTION = '" + uuid_execution + "'"
    df = pd.read_sql(query, engine)
    return df

def extract_permissions():
    engine = create_engine(f'mysql+mysqlconnector://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}/{db_config["database"]}')
    query = "SELECT HASH, APP_NAME, Permissions FROM Permissions WHERE ID_EXECUTION = '" + uuid_execution + "'"
    df = pd.read_sql(query, engine)
    return df

def collect_data_script(uuid_execution):

    # Extract data from the database
    report_table = extract_report()
    fail_counts_table = extract_total_fail_counts()
    logging_table = extract_logging()
    script_findings = extract_findings()
    permissions_table = extract_permissions()

    # Formula execution and data extraction
    database_utils.unify_suid_permissions(uuid_execution)

    # Obtain test list
    with open('config/methods_config.yml') as f:
        config = yaml.load(f, Loader=yaml.SafeLoader)

    tests_list = [value.upper() for values in config['tests'].values() for value in values]

    value = formula.calculate_formula(0.01, 0.01, tests_list, uuid_execution)
    formatted_value = '{:.4f}'.format(value).replace('.', ',')

    if export_csv():
        report_table.to_csv(path_export_csv + '/App_Analysis_Report_' + actual_timestamp + '.csv', index=False)
        fail_counts_table.to_csv(path_export_csv + '/App_Analysis_Total_Fail_Counts_' + actual_timestamp + '.csv', index=False)
        logging_table.to_csv(path_export_csv + '/App_Analysis_Logging_' + actual_timestamp + '.csv', index=False)
        script_findings.to_csv(path_export_csv + '/App_Analysis_Findings_' + actual_timestamp + '.csv', index=False)
        permissions_table.to_csv(path_export_csv + '/App_Analysis_Permissions_' + actual_timestamp + '.csv', index=False)
        data_formula = [['Formula'], [formatted_value]]
        df = pd.DataFrame(data_formula)
        # Export the DataFrame to a CSV file with no index and no header
        df.to_csv(path_export_csv + '/App_Analysis_Formula_' + actual_timestamp + '.csv', index=False, header=False)

    else:
        # Create a new Excel workbook
        workbook = Workbook()
        report_sheet = workbook.create_sheet(title='Report')
        findings_sheet = workbook.create_sheet(title='Findings')
        fail_counts_sheet = workbook.create_sheet(title='Total_Fail_Counts')
        formula_sheet = workbook.create_sheet(title='Formula')
        logging_sheet = workbook.create_sheet(title='Logging')
        permissions_sheet = workbook.create_sheet(title='Permissions')

        # Convert DataFrames to rows and write to the corresponding sheets
        for row in dataframe_to_rows(report_table, index=False, header=True):
            report_sheet.append(row)

        for row in dataframe_to_rows(script_findings, index=False, header=True):
            findings_sheet.append(row)

        for row in dataframe_to_rows(fail_counts_table, index=False, header=True):
            fail_counts_sheet.append(row)

        for row in dataframe_to_rows(permissions_table, index=False, header=True):
            permissions_sheet.append(row)

        for row in dataframe_to_rows(logging_table, index=False, header=True):
            logging_sheet.append(row)

        formula_sheet.append(["Formula Value"])
        formula_sheet.append([value])

        # Remove the default empty sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        # Save the workbook to a file
        workbook.save(path_export_csv + '/App_Analysis_' + actual_timestamp + '.xlsx')

def collect_data_semgrep():
    report_table = extract_report()
    finding_table = extract_findings()

    if export_csv():
        report_table.to_csv(path_export_csv + '/App_Analysis_Report_' + actual_timestamp + '.csv', index=False)
        finding_table.to_csv(path_export_csv + '/App_Analysis_Findings_' + actual_timestamp + '.csv', index=False)

    else:
        # Create a new Excel workbook
        workbook = Workbook()
        report_sheet = workbook.create_sheet(title='Report')
        report_findings = workbook.create_sheet(title='Findings')

        # Convert DataFrames to rows and write to the corresponding sheets
        for row in dataframe_to_rows(report_table, index=False, header=True):
            report_sheet.append(row)

        # Convert DataFrames to rows and write to the corresponding sheets
        for row in dataframe_to_rows(finding_table, index=False, header=True):
            report_findings.append(row)

        # Remove the default empty sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        # Save the workbook to a file
        workbook.save(path_export_csv + '/App_Analysis_' + actual_timestamp + '.xlsx')

if __name__ == "__main__":

    # Verify if some apps has been scanned with timestamp
    scanned_apks = database_utils.get_scanned_apks(uuid_execution)

    if scanned_apks == 0:
        print('Could not export data because there were no applications to scan.')
    else:

        use_script = not use_semgrep()

        if not os.path.exists(path_export_csv):
            try:
                os.mkdir(path_export_csv)
            except Exception as e:
                print(f"Error creating {path_export_csv} folder: {e}")

        if not os.path.exists(path_export_csv):
            try:
                os.mkdir(path_export_csv)
            except Exception as e:
                print(f"Error creating {path_export_csv} folder: {e}")

        if use_script:
            collect_data_script(uuid_execution)
        else:
            collect_data_semgrep()
