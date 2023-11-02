import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import mysql.connector  # Import MySQL connector library

import formula
import database_utils

# Replace with your actual database connection details
db_config = {
    "host": "localhost",
    "user": "jmsl",
    "password": "mysqlpassword",
    "database": "automated_MASA"
}

# Function to extract data from the database for Report table
def extract_data_table1():
    conn = mysql.connector.connect(**db_config)
    query = "SELECT * FROM Report"
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# Function to extract data from the database for Total Fail Counts table
def extract_data_table2():
    conn = mysql.connector.connect(**db_config)
    query = "SELECT * FROM Total_Fail_Counts"
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# Function to extract data from the database for Logging table
def extract_data_table4():
    conn = mysql.connector.connect(**db_config)
    query = "SELECT * FROM Logging"
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# Extract data from the database
df_table1 = extract_data_table1()
df_table2 = extract_data_table2()
df_table4 = extract_data_table4()

# Create a new Excel workbook
workbook = Workbook()
sheet1 = workbook.create_sheet(title='Report')
sheet2 = workbook.create_sheet(title='Total_Fail_Counts')
sheet3 = workbook.create_sheet(title='Formula')
sheet4 = workbook.create_sheet(title='Logging')

# Convert DataFrames to rows and write to the corresponding sheets
for row in dataframe_to_rows(df_table1, index=False, header=True):
    sheet1.append(row)

for row in dataframe_to_rows(df_table2, index=False, header=True):
    sheet2.append(row)

for row in dataframe_to_rows(df_table4, index=False, header=True):
    sheet4.append(row)



# Formula execution and data extraction

database_utils.unify_suid_permissions()
value = formula.calculate_formula(0.01, 0.01)


sheet3.append(["Formula Value"])
sheet3.append([value])

# Remove the default empty sheet created by openpyxl
if 'Sheet' in workbook.sheetnames:
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

# Save the workbook to a file
workbook.save('Preload_Analysis.xlsx')